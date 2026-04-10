"""Extract product data from Excel files (xlsx/xls).

Handles diverse vendor price list formats by:
1. Auto-detecting header rows via keyword matching
2. Mapping columns to product fields
3. Extracting vendor info from header area
4. Processing multiple sheets (each may be a product category)
"""

from __future__ import annotations

import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from wechat_automation.models import WeChatProduct

logger = logging.getLogger(__name__)

# Column header keywords -> field mapping (case-insensitive, bilingual)
_COLUMN_KEYWORDS: dict[str, list[str]] = {
    "product_name": ["product", "品名", "名称", "description", "item", "name", "features", "specification"],
    "sku": ["model", "型号", "code", "sku", "no.", "序号", "item no"],
    "dimensions": ["size", "尺寸", "dimension", "规格"],
    "unit_price": ["unit price", "price", "价格", "单价", "rmb", "usd", "报价"],
    "material": ["material", "材质", "fabric", "材料"],
    "color": ["color", "颜色", "colour"],
    "moq": ["moq", "起订", "minimum", "qty", "数量"],
    "weight": ["weight", "重量", "kg", "净重"],
    "description": ["remark", "备注", "notes", "说明", "description"],
}

# Keywords that indicate a header row
_HEADER_INDICATORS = {
    "product", "model", "price", "size", "item", "description", "specification",
    "品名", "型号", "价格", "尺寸", "规格", "名称", "单价", "材质",
    "no.", "image", "picture", "unit", "total", "qty", "code",
}

# Currency detection patterns
_CURRENCY_PATTERNS = [
    (re.compile(r"\$|usd|USD|美元", re.IGNORECASE), "USD"),
    (re.compile(r"rmb|RMB|¥|人民币|元", re.IGNORECASE), "CNY"),
    (re.compile(r"€|eur|EUR", re.IGNORECASE), "EUR"),
    (re.compile(r"฿|thb|THB|บาท", re.IGNORECASE), "THB"),
]


def extract_products_from_excel(
    filepath: str | Path,
    source_file_id: str = "",
    vendor_id: str = "",
    vendor_name: str = "",
) -> list[WeChatProduct]:
    """Extract products from an Excel file.

    Returns a list of WeChatProduct objects.
    """
    path = Path(filepath)
    ext = path.suffix.lower()

    if ext == ".xlsx":
        return _extract_xlsx(path, source_file_id, vendor_id, vendor_name)
    elif ext == ".xls":
        return _extract_xls(path, source_file_id, vendor_id, vendor_name)
    else:
        logger.warning("Unsupported extension: %s", ext)
        return []


def _extract_xlsx(
    path: Path, source_file_id: str, vendor_id: str, vendor_name: str,
) -> list[WeChatProduct]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    products: list[WeChatProduct] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        sheet_products = _extract_from_rows(
            rows=rows,
            sheet_name=sheet_name,
            source_file_id=source_file_id,
            source_filename=path.name,
            vendor_id=vendor_id,
            vendor_name=vendor_name,
        )
        products.extend(sheet_products)

    wb.close()
    logger.info("Extracted %d products from %s (%d sheets)", len(products), path.name, len(wb.sheetnames))
    return products


def _extract_xls(
    path: Path, source_file_id: str, vendor_id: str, vendor_name: str,
) -> list[WeChatProduct]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    products: list[WeChatProduct] = []

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        rows = []
        for row_idx in range(ws.nrows):
            rows.append(tuple(ws.cell_value(row_idx, col) for col in range(ws.ncols)))

        sheet_products = _extract_from_rows(
            rows=rows,
            sheet_name=ws.name,
            source_file_id=source_file_id,
            source_filename=path.name,
            vendor_id=vendor_id,
            vendor_name=vendor_name,
        )
        products.extend(sheet_products)

    logger.info("Extracted %d products from %s (%d sheets)", len(products), path.name, wb.nsheets)
    return products


def _extract_from_rows(
    rows: list[tuple],
    sheet_name: str,
    source_file_id: str,
    source_filename: str,
    vendor_id: str,
    vendor_name: str,
) -> list[WeChatProduct]:
    """Core extraction logic: find headers, map columns, extract data rows."""
    if len(rows) < 2:
        return []

    # Step 1: Find header row
    header_idx = _find_header_row(rows)
    if header_idx is None:
        logger.debug("No header row found in sheet '%s'", sheet_name)
        return []

    header = rows[header_idx]

    # Step 2: Map columns to fields
    col_map = _map_columns(header)
    if not col_map:
        return []

    # Step 3: Detect currency from header area
    currency = _detect_currency(rows[:header_idx + 1])

    # Step 4: Extract vendor name from header area if not provided
    if not vendor_name:
        vendor_name = _extract_vendor_from_header(rows[:header_idx])

    # Step 5: Use sheet name as category/subcategory
    category = _clean_sheet_name(sheet_name)

    # Step 6: Extract data rows
    products: list[WeChatProduct] = []
    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        product = _parse_data_row(
            row=row,
            col_map=col_map,
            row_num=row_idx + 1,
            sheet_name=sheet_name,
            category=category,
            currency=currency,
            source_file_id=source_file_id,
            source_filename=source_filename,
            vendor_id=vendor_id,
            vendor_name=vendor_name,
        )
        if product:
            products.append(product)

    return products


def _find_header_row(rows: list[tuple], max_scan: int = 15) -> int | None:
    """Find the header row by looking for rows with multiple known keywords."""
    best_idx = None
    best_score = 0

    for idx, row in enumerate(rows[:max_scan]):
        score = 0
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).lower().strip()
            for indicator in _HEADER_INDICATORS:
                if indicator in cell_str:
                    score += 1
                    break
        if score > best_score and score >= 2:
            best_score = score
            best_idx = idx

    return best_idx


def _map_columns(header: tuple) -> dict[str, int]:
    """Map column indices to product fields based on header keywords."""
    col_map: dict[str, int] = {}

    for col_idx, cell in enumerate(header):
        if cell is None:
            continue
        cell_str = str(cell).lower().strip()

        for field, keywords in _COLUMN_KEYWORDS.items():
            if field in col_map:
                continue
            for kw in keywords:
                if kw in cell_str:
                    col_map[field] = col_idx
                    break

    return col_map


def _detect_currency(rows: list[tuple]) -> str:
    """Detect currency from header area text."""
    for row in rows:
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell)
            for pattern, curr in _CURRENCY_PATTERNS:
                if pattern.search(cell_str):
                    return curr
    return "CNY"  # Default for Chinese vendors


def _extract_vendor_from_header(rows: list[tuple]) -> str:
    """Try to extract vendor name from the first few rows."""
    for row in rows[:5]:
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).strip()
            # Look for company name patterns
            if any(kw in cell_str for kw in ["Co.", "Ltd", "公司", "厂", "Corporation", "Inc"]):
                # Clean up the company name
                name = cell_str.split("\n")[0].strip()
                return name[:100]
    return ""


def _clean_sheet_name(name: str) -> str:
    """Clean sheet name for use as category."""
    # Remove common noise
    for remove in ["Sheet", "PRICE", "LIST", "NOTES", "价格", "报价"]:
        name = name.replace(remove, "")
    return name.strip(" -_").strip()


def _parse_data_row(
    row: tuple,
    col_map: dict[str, int],
    row_num: int,
    sheet_name: str,
    category: str,
    currency: str,
    source_file_id: str,
    source_filename: str,
    vendor_id: str,
    vendor_name: str,
) -> WeChatProduct | None:
    """Parse a single data row into a WeChatProduct."""

    def _get(field: str) -> str:
        idx = col_map.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    # Get product name — try product_name first, fall back to sku
    name = _get("product_name")
    sku = _get("sku")

    # Skip empty rows, section headers, and totals
    if not name and not sku:
        return None
    combined = (name + sku).lower()
    if any(skip in combined for skip in ["total", "subtotal", "合计", "小计", "备注"]):
        return None

    # Parse price
    price_str = _get("unit_price")
    unit_price = _parse_number(price_str)

    # Skip rows that look like section headers (no price and no sku)
    if not sku and unit_price == 0.0 and len(name) < 5:
        return None

    # Parse other fields
    dimensions = _get("dimensions")
    material = _get("material")
    color = _get("color")
    description = _get("description")
    moq = int(_parse_number(_get("moq"))) if _get("moq") else 0
    weight = _parse_number(_get("weight"))

    # Use name or sku as product name
    product_name = name if name else sku
    product_name_zh = ""
    # Split bilingual names if both Chinese and English present
    if re.search(r"[\u4e00-\u9fff]", product_name) and re.search(r"[a-zA-Z]", product_name):
        product_name_zh = product_name

    return WeChatProduct(
        product_name=product_name,
        product_name_zh=product_name_zh,
        source_file_id=source_file_id,
        source_filename=source_filename,
        source_page=row_num,
        sku=sku,
        description=description,
        category=category,
        material=material,
        dimensions=dimensions,
        weight_kg=weight,
        color=color,
        unit_price=unit_price,
        currency=currency,
        moq=moq,
        vendor_id=vendor_id,
        vendor_name=vendor_name,
        extraction_method="excel_parse",
        extraction_confidence=0.8,
    )


def _parse_number(s: str) -> float:
    """Parse a number from a string, handling various formats."""
    if not s:
        return 0.0
    # Already a number
    try:
        return float(s)
    except (ValueError, TypeError):
        pass
    # Remove currency symbols and whitespace
    cleaned = re.sub(r"[¥$€฿,\s]", "", str(s))
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0
